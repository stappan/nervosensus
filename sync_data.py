#!/usr/bin/env python3
"""
sync_data.py — Sync Excel → js/data.js

Reads the NPO cell type Excel file (.xlsx), parses the 'forNervoSensus' sheet,
and writes js/data.js with the four constants:
    DEFAULT_FAMILIES, DEFAULT_CELL_TYPES, DEFAULT_GENES, DEFAULT_SOURCES

Usage:
    python sync_data.py                     # auto-finds .xlsx in project dir
    python sync_data.py path/to/file.xlsx   # specify a file
"""

import json
import os
import sys
import glob

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required.  Install with:  pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SOURCE_COLORS = [
    '#667eea',  # CSA paper
    '#f59e0b',  # big DRG paper
    '#22c55e',  # Tavares-Ferreira
    '#ef4444',  # Yu et al.
    '#8b5cf6',  # Krauter et al.
    '#06b6d4',  # Qi et al.
    '#ec4899',  # Kupari et al.
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def safe_str(val):
    """Return trimmed string; treat None and 'none' (case-insensitive) as ''."""
    if val is None:
        return ''
    s = str(val).strip()
    if s.lower() == 'none':
        return ''
    return s


def find_xlsx(project_dir):
    """Find the first .xlsx file in the project directory."""
    candidates = glob.glob(os.path.join(project_dir, '*.xlsx'))
    if not candidates:
        return None
    # Prefer files with 'NPO' in the name
    for c in candidates:
        if 'NPO' in os.path.basename(c).upper():
            return c
    return candidates[0]


def read_excel(path):
    """Read the data sheet and return rows as list of dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Find the sheet (case-insensitive match, try multiple known names)
    sheet_name = None
    known_names = ['pnscellstonpo', 'fornervosensus']
    for name in wb.sheetnames:
        if name.lower().replace(' ', '') in known_names:
            sheet_name = name
            break

    if sheet_name is None:
        print(f"ERROR: No recognized sheet found. Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[sheet_name]
    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h else '' for h in row]
            continue
        if all(c is None for c in row):
            continue
        rows.append(dict(zip(headers, row)))
    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Main parsing logic — mirrors parseNPOData() in js/app.js lines 116-367
# ---------------------------------------------------------------------------

def parse_npo_data(rows):
    """
    Replicate the JavaScript parseNPOData() function.
    Returns (families, cell_types, genes, sources).
    """
    # ------------------------------------------------------------------
    # First pass: collect all data by neuron
    # ------------------------------------------------------------------
    neurons = {}  # neuron_id -> list of property dicts
    neuron_npokb = {}  # neuron_id -> npokb:Id CURIE
    skipped_dont_add = 0

    for row in rows:
        neuron_id = safe_str(row.get('Neuron ID'))
        npo_prop = safe_str(row.get('NPO Property'))
        value_label = row.get('Property Value Label')
        value_iri = row.get('NPO Property Value IRI')
        union_num = row.get('Union set number') or row.get('')
        nest_num = row.get('Nest intersection number')
        npokb_id = safe_str(row.get('npokb ID'))
        proposed_action = safe_str(row.get('Proposed action'))

        if not neuron_id:
            continue

        if proposed_action.lower() == "don't add":
            skipped_dont_add += 1
            continue

        val = safe_str(value_label)
        iri = safe_str(value_iri)

        if npokb_id and neuron_id:
            neuron_npokb[neuron_id] = npokb_id

        if neuron_id not in neurons:
            neurons[neuron_id] = []

        neurons[neuron_id].append({
            'npo': npo_prop,
            'union': union_num,
            'nest': nest_num,
            'value': val,
            'iri': iri,
        })

    print(f"  Rows skipped (don't add): {skipped_dont_add}")

    # ------------------------------------------------------------------
    # Build entity (rdfs:label) → npokb:Id mapping
    # ------------------------------------------------------------------
    entity_to_npokb = {}
    for nid, props in neurons.items():
        npokb_id = neuron_npokb.get(nid, '')
        for p in props:
            if p['npo'] == 'rdfs:label' and p['value'] and npokb_id:
                entity_to_npokb[p['value']] = npokb_id

    # ------------------------------------------------------------------
    # Build master cell npokb:Id set and prefLabel mapping
    # ------------------------------------------------------------------
    master_npokb_ids = set()
    master_npokb_to_pref = {}  # npokb:Id -> prefLabel

    for nid, props in neurons.items():
        if 'master' not in nid.lower():
            continue
        npokb_id = neuron_npokb.get(nid, '')
        if npokb_id:
            master_npokb_ids.add(npokb_id)
        pref = ''
        for p in props:
            if p['npo'] == 'skos:prefLabel':
                pref = p['value']
        if pref and npokb_id:
            master_npokb_to_pref[npokb_id] = pref

    # ------------------------------------------------------------------
    # Collect sources
    # ------------------------------------------------------------------
    sources = {}  # url -> { label, cells[] }

    for nid, props in neurons.items():
        if 'master' in nid.lower():
            continue
        for p in props:
            if p['npo'] == 'ilxtr:literatureCitation' and p['iri']:
                if p['iri'] not in sources:
                    sources[p['iri']] = {'label': p['value'] or p['iri'], 'cells': []}
                sources[p['iri']]['cells'].append(nid)

    source_color_map = {}
    for i, (url, data) in enumerate(sources.items()):
        source_color_map[url] = {
            'color': SOURCE_COLORS[i % len(SOURCE_COLORS)],
            'label': data['label'],
            'count': len(data['cells']),
        }

    # ------------------------------------------------------------------
    # Process cells
    # ------------------------------------------------------------------
    cell_types = []
    gene_map = {}  # base_name -> { display, cells[] }

    for nid, props in neurons.items():
        if 'master' in nid.lower():
            continue

        # Find source for this cell
        source_url = ''
        source_label = ''
        for p in props:
            if p['npo'] == 'ilxtr:literatureCitation' and p['iri']:
                source_url = p['iri']
                source_label = p['value'] or p['iri']

        sc = source_color_map.get(source_url, {})
        source_color = sc.get('color', '#667eea')

        ct = {
            'id': neuron_npokb.get(nid, ''),
            'entity': '',
            'preferredLabel': '',
            'species': 'unknown',
            'circuitRole': 'sensory',
            'neurotransmitter': '',
            'somaLocation': '',
            'somaLocations': [],
            'sensoryTerminalLocations': [],
            'axonTerminalLocations': [],
            'sourceNomenclature': source_url,
            'sourceNomenclatureLabel': source_label,
            'sourceColor': source_color,
            'markerGenes': [],
            'geneExpressionString': '',
            'fiberTypeString': '',
            'fiberTypeStringAbbrev': '',
            'physiologyString': '',
            'physiologyStringAbbrev': '',
            'creLine': '',
            'color': source_color,
            'masterLabel': '',
            'relatedCells': [],
            'mapsTo': [],
            'assertedSubclassOf': [],
            'geneBaseNames': [],
            'alertNotes': [],
            'curatorNotes': [],
            'sourceData': [],
            'localLabel': '',
            'clusterAttributes': {
                'cold_sensitive': False,
                'heat_sensitive': False,
                'mechanosensitive_ltm': False,
                'mechanosensitive_htm': False,
                'proprioceptive': False,
                'rapidly_adapting': False,
                'slowly_adapting': False,
                'fiber_a_beta': False,
                'fiber_a_delta': False,
                'fiber_c': False,
                'species_mouse': False,
                'species_human': False,
                'species_macaque': False,
                'species_guinea_pig': False,
                'soma_drg': False,
                'soma_tg': False,
            },
        }

        gene_items = []
        threshold_items = []
        adaptation_items = []
        functional_items = []
        axon_items = []

        for p in props:
            npo = p['npo']
            val = p['value']
            iri = p['iri']

            if npo == 'rdfs:label':
                ct['entity'] = val

            elif npo == 'skos:prefLabel':
                ct['preferredLabel'] = val

            elif npo == 'ilxtr:localLabel':
                ct['localLabel'] = val

            elif npo == 'ilxtr:hasInstanceInTaxon':
                ct['species'] = val.lower()
                sp_key = 'species_' + val.lower().replace(' ', '_')
                if sp_key in ct['clusterAttributes']:
                    ct['clusterAttributes'][sp_key] = True

            elif npo == 'ilxtr:hasCircuitRolePhenotype':
                ct['circuitRole'] = val

            elif npo == 'ilxtr:hasNeurotransmitterPhenotype':
                ct['neurotransmitter'] = val.strip()

            elif npo == 'ilxtr:hasSomaLocatedIn' and val:
                if val not in ct['somaLocations']:
                    ct['somaLocations'].append(val)
                if 'dorsal root' in val.lower():
                    ct['clusterAttributes']['soma_drg'] = True
                if 'trigeminal' in val.lower():
                    ct['clusterAttributes']['soma_tg'] = True

            elif npo == 'ilxtr:hasAxonSensoryTerminalLocatedIn' and val:
                if val not in ct['sensoryTerminalLocations']:
                    ct['sensoryTerminalLocations'].append(val)

            elif npo == 'ilxtr:hasAxonTerminalLocatedIn' and val:
                if val not in ct['axonTerminalLocations']:
                    ct['axonTerminalLocations'].append(val)

            elif npo == 'ilxtr:hasDriverExpressionConstitutivePhenotype':
                ct['creLine'] = val

            elif npo in ('TEMP:mapsTo', 'ilxtr:mapsTo'):
                target_id = iri if iri.startswith('npokb:') else entity_to_npokb.get(val, '')
                ct['mapsTo'].append({'label': val, 'iri': iri, 'id': target_id})

            elif npo in ('ilxtr:hasExpressionPhenotype',
                         'ilxtr:hasNucleicAcidExpressionPhenotype') and val:
                parts = val.split(' ', 1)
                nm = parts[0]
                exp = parts[1] if len(parts) > 1 else ''
                display = nm + '^' + exp if exp else nm
                gene_items.append({'union': p['union'], 'nest': p['nest'], 'display': display})
                ct['markerGenes'].append({'name': nm, 'uri': iri or '', 'expression': exp})
                base = nm.lower()
                if base not in ct['geneBaseNames']:
                    ct['geneBaseNames'].append(base)
                if base not in gene_map:
                    gene_map[base] = {'display': nm, 'cells': []}
                gene_map[base]['cells'].append(ct['preferredLabel'])

            elif npo == 'ilxtr:hasAxonPhenotype' and val:
                axon_items.append({'union': p['union'], 'nest': p['nest'], 'display': val})
                vl = val.lower()
                if 'beta' in vl or '(beta)' in vl:
                    ct['clusterAttributes']['fiber_a_beta'] = True
                if 'delta' in vl or '(delta)' in vl:
                    ct['clusterAttributes']['fiber_a_delta'] = True
                if 'type c' in vl:
                    ct['clusterAttributes']['fiber_c'] = True

            elif npo in ('ilxtr:hasThresholdPhenotype',
                         'ilxtr:hasPredictedThresholdPhenotype') and val:
                threshold_items.append({'union': p['union'], 'nest': p['nest'], 'display': val})
                vl = val.lower()
                if 'ltm' in vl or 'low-threshold' in vl:
                    ct['clusterAttributes']['mechanosensitive_ltm'] = True
                if 'htm' in vl or 'high-threshold' in vl:
                    ct['clusterAttributes']['mechanosensitive_htm'] = True

            elif npo == 'ilxtr:hasAdaptationPhenotype' and val:
                adaptation_items.append({'union': p['union'], 'nest': p['nest'], 'display': val})
                vl = val.lower()
                if 'rapidly' in vl or '(ra)' in vl:
                    ct['clusterAttributes']['rapidly_adapting'] = True
                if 'slowly' in vl or 'sa1' in vl:
                    ct['clusterAttributes']['slowly_adapting'] = True

            elif npo == 'ilxtr:hasFunctionalPhenotype' and val:
                functional_items.append({'union': p['union'], 'nest': p['nest'], 'display': val})
                vl = val.lower()
                if 'cold' in vl:
                    ct['clusterAttributes']['cold_sensitive'] = True
                if 'heat' in vl:
                    ct['clusterAttributes']['heat_sensitive'] = True
                if 'proprioceptive' in vl:
                    ct['clusterAttributes']['proprioceptive'] = True

            elif npo == 'alertNote' and val:
                ct['alertNotes'].append(val)

            elif npo == 'curatorNote' and val:
                ct['curatorNotes'].append(val)

            elif npo == 'ilxtr:dataCitation' and (val or iri):
                ct['sourceData'].append({'label': val or iri, 'uri': iri or ''})

            elif npo in ('TEMP:subClassOf', 'TEMP:assertedSubClassOf'):
                target_id = iri if iri.startswith('npokb:') else entity_to_npokb.get(val, '')
                ct['assertedSubclassOf'].append({'label': val, 'iri': iri, 'id': target_id})

        # Build derived strings
        ct['geneExpressionString'] = ' + '.join(g['display'] for g in gene_items)
        ct['fiberTypeString'] = ' + '.join(a['display'] for a in axon_items)
        ct['fiberTypeStringAbbrev'] = ct['fiberTypeString']

        all_phys = threshold_items + adaptation_items + functional_items
        ct['physiologyString'] = ' + '.join(p['display'] for p in all_phys)
        ct['physiologyStringAbbrev'] = ct['physiologyString']

        if ct['somaLocations']:
            ct['somaLocation'] = ct['somaLocations'][0]

        # Only include cells that have a preferredLabel
        if ct['preferredLabel']:
            cell_types.append(ct)

    # ------------------------------------------------------------------
    # Derive masterLabel and relatedCells from assertedSubclassOf triples
    # ------------------------------------------------------------------
    master_to_children = {}
    for ct in cell_types:
        for rel in ct['assertedSubclassOf']:
            if rel['id'] in master_npokb_ids:
                ct['masterLabel'] = master_npokb_to_pref.get(rel['id'], rel['label'])
                if ct['masterLabel'] not in master_to_children:
                    master_to_children[ct['masterLabel']] = []
                master_to_children[ct['masterLabel']].append({
                    'id': ct['id'],
                    'label': ct['preferredLabel'],
                    'species': ct['species'],
                })
                break

    for ct in cell_types:
        ml = ct['masterLabel']
        if ml and ml in master_to_children:
            ct['relatedCells'] = [
                {'id': sib['id'], 'label': sib['label'], 'species': sib['species']}
                for sib in master_to_children[ml]
                if sib['label'] != ct['preferredLabel']
            ]

    # ------------------------------------------------------------------
    # Build genes list
    # ------------------------------------------------------------------
    genes = []
    for base, data in gene_map.items():
        genes.append({
            'id': 'gene_' + base,
            'base': base,
            'display': data['display'],
            'cells': list(dict.fromkeys(data['cells'])),  # deduplicate, preserve order
        })

    # ------------------------------------------------------------------
    # Build families from master cells (preserving order from Excel)
    # ------------------------------------------------------------------
    # Build preferredLabel → npokb:Id mapping for child lookups
    pref_to_npokb = {}
    for ct in cell_types:
        if ct['preferredLabel'] and ct['id']:
            pref_to_npokb[ct['preferredLabel']] = ct['id']

    families = []
    for nid, props in neurons.items():
        if 'master' not in nid.lower():
            continue
        pref_label = ''
        for p in props:
            if p['npo'] == 'skos:prefLabel':
                pref_label = p['value']
        if not pref_label:
            continue
        children = [{'id': c['id'], 'label': c['label']} for c in master_to_children.get(pref_label, [])]
        families.append({
            'id': neuron_npokb.get(nid, ''),
            'name': pref_label,
            'children': children,
        })

    return families, cell_types, genes, source_color_map


# ---------------------------------------------------------------------------
# Write js/data.js
# ---------------------------------------------------------------------------

def write_data_js(out_path, families, cell_types, genes, source_color_map):
    """Write the four constants to js/data.js."""

    # Convert Python booleans to JS-compatible JSON (true/false)
    families_json = json.dumps(families, ensure_ascii=False, separators=(',', ':'))
    cells_json = json.dumps(cell_types, ensure_ascii=False, separators=(',', ':'))
    genes_json = json.dumps(genes, ensure_ascii=False, separators=(',', ':'))
    sources_json = json.dumps(source_color_map, ensure_ascii=False, separators=(',', ':'))

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(f'const DEFAULT_FAMILIES = {families_json};\n')
        f.write(f'const DEFAULT_CELL_TYPES = {cells_json};\n')
        f.write(f'const DEFAULT_GENES = {genes_json};\n')
        f.write('\n\n')
        f.write(f'const DEFAULT_SOURCES = {sources_json};\n')

    print(f"Wrote {out_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    project_dir = os.path.dirname(os.path.abspath(__file__))

    # Determine Excel file path
    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]
    else:
        xlsx_path = find_xlsx(project_dir)

    if not xlsx_path or not os.path.isfile(xlsx_path):
        print("ERROR: No .xlsx file found. Provide a path:  python sync_data.py path/to/file.xlsx")
        sys.exit(1)

    print(f"Reading: {os.path.basename(xlsx_path)}")

    # Read and parse
    rows = read_excel(xlsx_path)
    print(f"  Rows read: {len(rows)}")

    families, cell_types, genes, source_color_map = parse_npo_data(rows)

    # Summary
    print(f"\n--- Summary ---")
    print(f"  Families:    {len(families)}")
    print(f"  Cell types:  {len(cell_types)}")
    print(f"  Genes:       {len(genes)}")
    print(f"  Sources:     {len(source_color_map)}")
    for url, sc in source_color_map.items():
        print(f"    {sc['label']}: {sc['count']} cells  ({sc['color']})")

    # Warnings
    cells_without_master = [ct['preferredLabel'] for ct in cell_types if not ct['masterLabel']]
    if cells_without_master:
        print(f"\n  WARNING: {len(cells_without_master)} cells have no masterLabel (no family):")
        for lbl in cells_without_master:
            print(f"    - {lbl}")

    cells_without_source = [ct['preferredLabel'] for ct in cell_types if not ct['sourceNomenclature']]
    if cells_without_source:
        print(f"\n  WARNING: {len(cells_without_source)} cells have no source citation:")
        for lbl in cells_without_source:
            print(f"    - {lbl}")

    # Write output
    out_path = os.path.join(project_dir, 'js', 'data.js')
    write_data_js(out_path, families, cell_types, genes, source_color_map)

    print(f"\nDone! {len(families)} families, {len(cell_types)} cells, "
          f"{len(genes)} genes, {len(source_color_map)} sources.")


if __name__ == '__main__':
    main()
